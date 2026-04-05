"""
Adds a second tab to the Apollo enrichment workbook:
'Apollo Tab 2 — Name Only (No Address)' containing the 1,393
scheme-address victims where Wincham replaced every address field.
Apollo can still match these on company name + director name alone.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, re

BASE    = r"C:\DAD\UK_Lanzarote_Repatriation\Wincham Legal Case UK Type Class Action"
MASTER  = os.path.join(BASE, "Wincham_Victim_Database_MASTER_v2.xlsx")
APOLLO  = os.path.join(BASE, "Wincham_Victims_Apollo_Enriched.xlsx")
CSV_OUT = os.path.join(BASE, "apollo_import_no_address_victims.csv")

# ── Parse name ────────────────────────────────────────────────────────────
def parse_name(raw):
    if not isinstance(raw, str) or not raw.strip():
        return "", ""
    raw = raw.strip()
    if "," in raw:
        parts = raw.split(",", 1)
        last  = parts[0].strip().title()
        first = parts[1].strip().title().split()[0] if parts[1].strip() else ""
    else:
        words = raw.split()
        last  = words[-1].title() if words else ""
        first = words[0].title()  if len(words) > 1 else ""
    return first, last

def entity_label(addr):
    if not isinstance(addr, str): return "Unknown Wincham Entity"
    if "4AA"          in addr: return "Adrem Accounting Ltd (CW12 4AA)"
    if "Wincham House" in addr: return "Wincham House"
    if "4TR"          in addr: return "Wincham Accountants (CW12 4TR)"
    if "Oak Drive"    in addr: return "Malcolm Roach Personal (ST7)"
    if "Woodberry"    in addr: return "Duke/Black Nominee (N12)"
    if "M25"          in addr or "M7 4AS" in addr: return "Jacobs/Heiman Nominee (M25)"
    return "Unknown Wincham Entity"

# ── Load scheme address sheet ─────────────────────────────────────────────
print("\n=== Adding Apollo Tab 2 — Name-Only Scheme Victims ===\n")
df = pd.read_excel(MASTER, sheet_name="📋 Scheme Address Entries", dtype=str).fillna("")

df[["First Name", "Last Name"]] = df["Director Name"].apply(
    lambda x: pd.Series(parse_name(x))
)

addr_col = "Director Address" if "Director Address" in df.columns else "Address Type"

tab2 = pd.DataFrame({
    "First Name":          df["First Name"],
    "Last Name":           df["Last Name"],
    "Title":               "",
    "Company":             df["Company Name"],
    "Company Number (CH)": df["Company Number"],
    "Company Status":      df["Status"],
    # Address intentionally blank — Wincham replaced it
    "Street":              "",
    "City":                "",
    "State":               "England",
    "Postal Code":         "",
    "Country":             "United Kingdom",
    # Apollo fills these
    "Email":               "",
    "Email Status":        "",
    "Phone":               "",
    "LinkedIn URL":        "",
    "Apollo ID":           "",
    # Context
    "Wincham Entity":      df[addr_col].apply(entity_label) if addr_col in df.columns else df.get("Scheme Address Used",""),
    "Note":                "Address suppressed by Wincham — enrich via Apollo name+company match",
    "Source File":         df.get("Source", df.get("Source File", "")),
    "Full Director Name":  df["Director Name"],
    "Wincham Address Used":df[addr_col] if addr_col in df.columns else "",
})

# Remove blank names
tab2 = tab2[
    (tab2["First Name"].str.strip() != "") |
    (tab2["Last Name"].str.strip()  != "")
].drop_duplicates().sort_values("Last Name")

tab2.to_csv(CSV_OUT, index=False, encoding="utf-8-sig")
print(f"  ✓ CSV saved: {CSV_OUT}  ({len(tab2):,} rows)")

# ── Open existing Apollo workbook and add tab ─────────────────────────────
wb = load_workbook(APOLLO)

# Remove existing Tab 2 if re-running
if "Apollo Tab 2 — Name Only" in wb.sheetnames:
    del wb["Apollo Tab 2 — Name Only"]

# Create new sheet
ws2 = wb.create_sheet("Apollo Tab 2 — Name Only")

# Write headers + data
for ci, col in enumerate(tab2.columns, 1):
    ws2.cell(row=1, column=ci, value=col)
for ri, row in enumerate(tab2.itertuples(index=False), 2):
    for ci, val in enumerate(row, 1):
        ws2.cell(row=ri, column=ci, value=val if val != "" else None)

# ── Styling ───────────────────────────────────────────────────────────────
ORANGE = PatternFill("solid", fgColor="7A3B00")
EMPTY  = PatternFill("solid", fgColor="FFF9C4")
ALT    = PatternFill("solid", fgColor="FFF3E0")
H_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
D_FONT = Font(name="Calibri", size=10)
E_FONT = Font(name="Calibri", size=10, italic=True, color="999999")
THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

APOLLO_COLS = {"Email", "Email Status", "Phone", "LinkedIn URL", "Apollo ID"}

for ci, col in enumerate(tab2.columns, 1):
    c = ws2.cell(row=1, column=ci)
    c.font      = H_FONT
    c.fill      = EMPTY if col in APOLLO_COLS else ORANGE
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER

for ri in range(2, len(tab2) + 2):
    fill = ALT if ri % 2 == 0 else None
    for ci in range(1, len(tab2.columns) + 1):
        c        = ws2.cell(row=ri, column=ci)
        col_name = tab2.columns[ci - 1]
        c.border    = BORDER
        c.alignment = Alignment(vertical="center")
        if col_name in APOLLO_COLS:
            c.font = E_FONT
            c.fill = EMPTY
        else:
            c.font = D_FONT
            if fill: c.fill = fill

col_widths = {
    "First Name": 15, "Last Name": 20, "Title": 12, "Company": 40,
    "Company Number (CH)": 18, "Company Status": 14, "Street": 12,
    "City": 12, "State": 10, "Postal Code": 10, "Country": 12,
    "Email": 30, "Email Status": 14, "Phone": 16, "LinkedIn URL": 32,
    "Apollo ID": 18, "Wincham Entity": 35, "Note": 45, "Source File": 20,
    "Full Director Name": 28, "Wincham Address Used": 40,
}
for ci, col in enumerate(tab2.columns, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 16)

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

# ── Update HOW TO USE APOLLO sheet with Tab 2 instructions ───────────────
if "HOW TO USE APOLLO" in wb.sheetnames:
    ws_inst = wb["HOW TO USE APOLLO"]
    next_row = ws_inst.max_row + 2
    GOLD = PatternFill("solid", fgColor="856404")
    ws_inst.cell(row=next_row, column=1, value="── Tab 2 Specific Steps ──").font = Font(bold=True, name="Calibri")
    steps_tab2 = [
        ("Tab 2 — Import separately", "Import 'apollo_import_no_address_victims.csv' as a SEPARATE Apollo list"),
        ("Tab 2 — No address matching", "Apollo will match on First Name + Last Name + Company Name only — lower match rate expected (~40-60%)"),
        ("Tab 2 — Verify carefully", "Cross-check any matched emails against company records before use — name-only matches carry more risk"),
        ("Tab 2 — Value proposition", "Even 40% match rate = ~557 additional contacts on top of the 801 verified — ~1,358 total enriched"),
    ]
    for i, (step, action) in enumerate(steps_tab2, next_row + 1):
        ws_inst.cell(row=i, column=1, value=step).font = Font(name="Calibri", size=10, bold=True)
        ws_inst.cell(row=i, column=2, value=action).font = Font(name="Calibri", size=10)
        for ci in range(1, 3):
            ws_inst.cell(row=i, column=ci).border = BORDER

wb.save(APOLLO)
print(f"  ✓ Apollo workbook updated: {APOLLO}")
print(f"\n  Tab 1 — Verified Victims (with home address) : 801 contacts")
print(f"  Tab 2 — Name-Only (Wincham address suppressed): {len(tab2):,} contacts")
print(f"  Combined potential at 40% Tab 2 match rate   : ~{801 + int(len(tab2)*0.4):,} enriched emails")
print(f"\n  ✅ Done. Import Tab 2 CSV separately into Apollo as a second list.")
