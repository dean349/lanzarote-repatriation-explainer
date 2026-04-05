"""
Wincham Victim Database — v2
Filters known scheme operators OUT of the victim sheet entirely.
Any name on the KNOWN_OPERATORS blocklist is moved to a
'⛔ DO NOT CONTACT' sheet with explanation.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

BASE = r"C:\DAD\UK_Lanzarote_Repatriation\Wincham Legal Case UK Type Class Action"
OUT  = os.path.join(BASE, "Wincham_Victim_Database_MASTER_v2.xlsx")

# ══════════════════════════════════════════════════════════════════════════════
# KNOWN OPERATOR BLOCKLIST
# These individuals are confirmed or highly probable Wincham/Adrem associates.
# ANY row containing these names must be excluded from the victim/claimant list.
# ══════════════════════════════════════════════════════════════════════════════
KNOWN_OPERATORS = {
    # Core Wincham principals
    "ROACH, Malcolm David":        "Core Wincham principal — Malcolm Roach (founder/director)",
    "ROACH, Mark Damion":          "Core Wincham principal — Mark Roach (son/co-director)",
    "IVARS, David":                "Wincham associate director — David Ivars",
    "VIVES IVARS, Jaime":          "Wincham associate director — Jaime Vives Ivars",
    "VIVES-IVARS, Jaime":          "Wincham associate director — Jaime Vives Ivars (alt spelling)",
    "HALLATT, Jon Antony":         "Wincham/Wincham House associate — Jon Hallatt",
    "JONES, Leonard Edward":       "Adrem Accounting director — Leonard Jones",
    "MATTHEWS, Robert Morton":     "Wincham Consultants director — Robert Matthews",
    "SPOONER, Paul Rodney":        "Wincham Consultants director — Paul Spooner",
    # Italian network linked to Wincham House era
    "TATA, Enzo":                  "Italian Wincham associate — Enzo Tata",
    "CACCIAMANI, Gabriele":        "Italian Wincham associate — Gabriele Cacciamani",
    "MACCARINI, Stefanella":       "Wincham House associate — Stefanella Maccarini",
    "FRISONE, Fulvio":             "Wincham House associate — Fulvio Frisone",
    "BALZANO, Federico":           "Wincham House associate — Federico Balzano",
    "PIRONE, Ciro":                "Wincham House associate — Ciro Pirone",
    "SALAMONE, Giuseppe":          "Wincham network associate — Giuseppe Salamone",
    "AIESI, Francesco":            "Wincham network associate — Francesco Aiesi",
    "BELLONI, Adriano":            "Wincham network associate — Adriano Belloni",
    "COMANDINI, Daniele":          "Wincham network associate — Daniele Comandini",
    "SCIMECA, Paolino":            "Wincham network associate — Paolino Scimeca",
    "TATA, Salvatore":             "Wincham network associate — Salvatore Tata",
    # Nominee directors identified across many companies
    "DUKE, Michael":               "Nominee director — Michael Duke (2 Woodberry Grove, Finchley)",
    "BLACK, Marion":               "Nominee director — Marion Black (2 Woodberry Grove, Finchley)",
    "JACOBS, Yomtov Eliezer":      "Nominee director — Yomtov Jacobs (Manchester M25)",
    "HEIMAN, Osker":               "Nominee director — Osker Heiman (Manchester M25)",
    # Other Wincham-linked recurring names
    "NAVARRO TUDELA, Maria Pilar": "Wincham associate — Maria Pilar Navarro Tudela",
    "SAFFELL, Oliver Henry James": "Wincham associate — Oliver Saffell",
    "GAITANOS, Yiannis":           "Wincham associate — Yiannis Gaitanos",
    "MALCOLM, John Quenton":       "Wincham-linked — John Quenton Malcolm",
    "NOUMOFF, Samuel Joseph":      "Wincham House associate — Samuel Noumoff",
    "WUEST DE WELLBERG, Frances Emmy": "Wincham House associate — Frances Wuest De Wellberg",
    "FARRUGIA, John":              "Wincham House associate — John Farrugia",
    "CIURCINA, Vittorio":          "Wincham House associate — Vittorio Ciurcina",
    "ROACH, Mark Damion":          "Core Wincham principal (duplicate key normalisation)",
    "KENWRIGHT, Lawrence":         "Wincham associate — Lawrence Kenwright",
    "FAUN, Stephen":               "Wincham/Corporate Funding associate — Stephen Faun",
    "AUSTEN, Paul James":          "Wincham/Corporate Funding associate — Paul Austen",
    "BIRCHALL, Timothy Martin":    "Corporate Funding/Wincham associate — Timothy Birchall",
    "COX, Lionel":                 "Wincham associate — Lionel Cox",
    "PATHA, Nikhila":              "Wincham nominee — Nikhila Patha",
    "PONS, Michael Nicholas":      "Wincham associate — Michael Pons",
    "CARRINGTON, Andrew Thomas":   "UK Asset Holdings associate — Andrew Carrington",
    "PODROJSKI, Sharon":           "Wincham-linked nominee — Sharon Podrojski",
    "EARDLEY, Roy":                "Wincham-linked — Roy Eardley",
    "ALCOCK, Stuart":              "Wincham network — Stuart Alcock",
    "ALLCOCK, Stuart":             "Wincham network — Stuart Allcock (alt spelling)",
    "KNIGHT, Michael Richard":     "Wincham network — Michael Knight",
    "NERI, Lewis":                 "Wincham network — Lewis Neri",
}

# Normalise blocklist keys to uppercase for matching
BLOCKLIST_UPPER = {k.upper(): v for k, v in KNOWN_OPERATORS.items()}

def is_operator(name):
    if not isinstance(name, str):
        return False
    return name.strip().upper() in BLOCKLIST_UPPER

def operator_reason(name):
    if not isinstance(name, str):
        return ""
    return BLOCKLIST_UPPER.get(name.strip().upper(), "")

# ── Wincham entity address signatures ─────────────────────────────────────
SCHEME_SIGS = [
    "CW12 4TR",  # catches all Wincham/Greenfield variants
    "CW12 4AA",  # Adrem
    "Wincham House",
    "22 Oak Drive, Scholar Green",
    "2 Woodberry Grove, Finchley",
    "Floor, Manchester, M25",
    "Bury New Road, Manchester, M25",
    "Leicester Road, Manchester, M7 4AS",
    "Prestwich, Manchester, M25",
]

def entity_label(addr):
    if not isinstance(addr, str):
        return "Unknown"
    for sig in SCHEME_SIGS:
        if sig.lower() in addr.lower():
            if "4AA" in sig:
                return "Adrem Accounting Ltd (CW12 4AA)"
            if "Wincham House" in sig:
                return "Wincham House"
            if "Oak Drive" in sig:
                return "Malcolm Roach Personal (ST7)"
            if "Woodberry Grove" in sig:
                return "Duke/Black Nominee (N12)"
            if "Manchester" in sig or "Prestwich" in sig or "Leicester Road" in sig:
                return "Jacobs/Heiman Nominee (M25)"
            return "Wincham Accountants (CW12 4TR)"
    return "Real Home Address"

# ── Load CSVs ──────────────────────────────────────────────────────────────
def load_leads(fname, source):
    df = pd.read_csv(os.path.join(BASE, fname), encoding="utf-8-sig", low_memory=False)
    df.columns = df.columns.str.strip()
    df = df.rename(columns={
        "Director Address (LEAD)":            "Director Address",
        "Director Correspondence Address (LEAD)": "Director Address",
        "Date of Creation":                   "Date Incorporated",
    })
    df["Source"] = source
    return df

wincham_df = load_leads("wincham_leads.csv",         "Wincham (CW12 4TR)")
adrem_df   = load_leads("adrem_leads.csv",           "Adrem (CW12 4TR)")
adrem_cw   = load_leads("adrem_cw12_4aa_leads.csv",  "Adrem (CW12 4AA)")

all_df = pd.concat([wincham_df, adrem_df, adrem_cw], ignore_index=True)

# Ensure columns
for col in ["Company Name","Company Number","Status","Date Incorporated","Director Name","Director Address","Source"]:
    if col not in all_df.columns:
        all_df[col] = ""

# ── Classify ───────────────────────────────────────────────────────────────
all_df["Address Type"]      = all_df["Director Address"].apply(entity_label)
all_df["Is Scheme Address"] = all_df["Address Type"] != "Real Home Address"
all_df["Is Operator"]       = all_df["Director Name"].apply(is_operator)
all_df["Operator Reason"]   = all_df["Director Name"].apply(operator_reason)

# ── Frequency analysis — flag high-frequency directors ───────────────────
freq = all_df.groupby("Director Name")["Company Name"].nunique().reset_index()
freq.columns = ["Director Name", "Companies Count"]
freq["High Frequency Flag"] = freq["Companies Count"].apply(
    lambda x: "⚠️ Appears in 5+ companies" if x >= 5 else ""
)
all_df = all_df.merge(freq, on="Director Name", how="left")

# ── Final victim list: real home address AND not a known operator ──────────
clean_victims = all_df[
    (~all_df["Is Scheme Address"]) &
    (~all_df["Is Operator"])
].copy()

# Operators (known blocklist hits)
operators_blocked = all_df[all_df["Is Operator"]].copy()

# Scheme-address rows that aren't on the blocklist (company clients whose address was replaced)
scheme_addr_rows = all_df[
    (all_df["Is Scheme Address"]) &
    (~all_df["Is Operator"])
].copy()

# ── Output columns ───────────────────────────────────────────────────────
VICTIM_COLS = [
    "Director Name",
    "Director Address",    # REAL home address
    "Company Name",
    "Company Number",
    "Status",
    "Date Incorporated",
    "Address Type",
    "Companies Count",
    "High Frequency Flag",
    "Source",
]

OPERATOR_COLS = [
    "Director Name",
    "Director Address",
    "Operator Reason",
    "Company Name",
    "Company Number",
    "Status",
    "Companies Count",
    "Source",
]

def safe_cols(df, cols):
    return df[[c for c in cols if c in df.columns]].drop_duplicates().sort_values(cols[0])

sheets = {
    "✅ VERIFIED VICTIMS":          safe_cols(clean_victims,     VICTIM_COLS),
    "⛔ DO NOT CONTACT (Operators)": safe_cols(operators_blocked, OPERATOR_COLS),
    "📋 Scheme Address Entries":    safe_cols(scheme_addr_rows,  VICTIM_COLS),
    "Wincham CW12 4TR":             safe_cols(all_df[all_df["Address Type"].str.contains("CW12 4TR", na=False)], VICTIM_COLS),
    "Adrem CW12 4AA":               safe_cols(all_df[all_df["Address Type"].str.contains("4AA", na=False)],      VICTIM_COLS),
    "Wincham House":                safe_cols(all_df[all_df["Address Type"].str.contains("Wincham House", na=False)], VICTIM_COLS),
    "Duke-Black Finchley":          safe_cols(all_df[all_df["Address Type"].str.contains("Duke", na=False)],     VICTIM_COLS),
    "Jacobs-Heiman Manchester":     safe_cols(all_df[all_df["Address Type"].str.contains("Jacobs", na=False)],   VICTIM_COLS),
}

# ── Write ─────────────────────────────────────────────────────────────────
print("\n=== Building Wincham Victim Database v2 (with Operator Blocklist) ===\n")

with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    for name, data in sheets.items():
        sheet_name = name[:31]
        data.to_excel(writer, sheet_name=sheet_name, index=False)

# Style
NAVY   = PatternFill("solid", fgColor="1F3864")
RED    = PatternFill("solid", fgColor="8B0000")
GREEN  = PatternFill("solid", fgColor="1A5C2A")
ORANGE = PatternFill("solid", fgColor="7A3B00")
TEAL   = PatternFill("solid", fgColor="1E6B7A")
PURPLE = PatternFill("solid", fgColor="4B3A8C")
DARK   = PatternFill("solid", fgColor="2D2D2D")
GREY   = PatternFill("solid", fgColor="555555")
ALT_G  = PatternFill("solid", fgColor="E8F5E9")
ALT_R  = PatternFill("solid", fgColor="FFEBEE")
ALT_B  = PatternFill("solid", fgColor="EBF3FB")

H_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
D_FONT = Font(name="Calibri", size=10)
THIN   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

sheet_fills = {
    "✅ VERIFIED VICTIMS":           (GREEN,  ALT_G),
    "⛔ DO NOT CONTACT (Operators)": (RED,    ALT_R),
    "📋 Scheme Address Entries":    (ORANGE, ALT_B),
    "Wincham CW12 4TR":             (NAVY,   ALT_B),
    "Adrem CW12 4AA":               (TEAL,   ALT_B),
    "Wincham House":                (PURPLE, ALT_B),
    "Duke-Black Finchley":          (DARK,   ALT_B),
    "Jacobs-Heiman Manchester":     (GREY,   ALT_B),
}

wb = load_workbook(OUT)

for sheet_name_full, (hf, af) in sheet_fills.items():
    sn = sheet_name_full[:31]
    if sn not in wb.sheetnames:
        continue
    ws = wb[sn]
    data = sheets[sheet_name_full]

    for ci in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=ci)
        c.font = H_FONT
        c.fill = hf
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for ri in range(2, ws.max_row + 1):
        fill = af if ri % 2 == 0 else None
        for ci in range(1, ws.max_column + 1):
            c = ws.cell(row=ri, column=ci)
            c.font = D_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical="center")
            if fill:
                c.fill = fill

    for ci in range(1, ws.max_column + 1):
        col_letter = get_column_letter(ci)
        vals = [ws.cell(row=r, column=ci).value or "" for r in range(1, min(ws.max_row+1, 200))]
        max_len = max((len(str(v)) for v in vals), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 70)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    print(f"  ✓ {sheet_name_full[:50]:50s}  {len(data):>5,} rows")

# Summary sheet
ws_sum = wb.create_sheet("📊 SUMMARY")
summary_rows = [
    ["Sheet", "Rows", "Action Required", "Notes"],
    ["✅ VERIFIED VICTIMS", len(sheets["✅ VERIFIED VICTIMS"]),
     "SAFE TO CONTACT", "Real home addresses, operators removed — primary claimant list"],
    ["⛔ DO NOT CONTACT (Operators)", len(sheets["⛔ DO NOT CONTACT (Operators)"]),
     "⛔ NEVER CONTACT", "Known Wincham/Adrem associates — contacting these is a serious risk"],
    ["📋 Scheme Address Entries", len(sheets["📋 Scheme Address Entries"]),
     "REVIEW BEFORE USE", "Address replaced by Wincham address — real address unknown, needs CH lookup"],
    ["Wincham CW12 4TR", len(sheets["Wincham CW12 4TR"]), "REFERENCE ONLY", "All entries routed through main Wincham CW12 4TR address"],
    ["Adrem CW12 4AA", len(sheets["Adrem CW12 4AA"]), "REFERENCE ONLY", "All entries routed through Adrem CW12 4AA"],
    ["Wincham House", len(sheets["Wincham House"]), "REFERENCE ONLY", "Early Wincham House entity"],
    ["Duke-Black Finchley", len(sheets["Duke-Black Finchley"]), "REFERENCE ONLY", "Nominee directors Michael Duke / Marion Black"],
    ["Jacobs-Heiman Manchester", len(sheets["Jacobs-Heiman Manchester"]), "REFERENCE ONLY", "Nominee directors Yomtov Jacobs / Osker Heiman"],
]

header_fills_sum = [GREEN, RED, ORANGE, NAVY, TEAL, PURPLE, DARK, GREY, NAVY]
for ri, row in enumerate(summary_rows, 1):
    for ci, val in enumerate(row, 1):
        c = ws_sum.cell(row=ri, column=ci, value=val)
        if ri == 1:
            c.font = H_FONT; c.fill = NAVY; c.alignment = Alignment(horizontal="center"); c.border = BORDER
        else:
            c.font = D_FONT; c.border = BORDER
            if ci == 1:
                c.fill = header_fills_sum[ri-2] if ri-2 < len(header_fills_sum) else NAVY
                c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

ws_sum.column_dimensions["A"].width = 38
ws_sum.column_dimensions["B"].width = 10
ws_sum.column_dimensions["C"].width = 22
ws_sum.column_dimensions["D"].width = 75
ws_sum.freeze_panes = "A2"

# Move summary to front
wb.move_sheet("📊 SUMMARY", offset=-len(wb.sheetnames)+1)
wb.save(OUT)

print(f"\n✅ Saved: {OUT}")
print(f"\n  KEY STATS:")
print(f"  Safe-to-contact verified victims : {len(sheets['✅ VERIFIED VICTIMS']):,}")
print(f"  DO NOT CONTACT (operators)       : {len(sheets['⛔ DO NOT CONTACT (Operators)']):,}")
print(f"  Scheme address (unknown real addr): {len(sheets['📋 Scheme Address Entries']):,}")
print(f"\n  Operator blocklist entries        : {len(KNOWN_OPERATORS)}")
