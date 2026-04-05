import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys

BASE_DIR = r"C:\DAD\UK_Lanzarote_Repatriation\Wincham Legal Case UK Type Class Action"

CSV_FILES = [
    "adrem_cw12_4aa_leads.csv",
    "adrem_leads.csv",
    "adrem_only_companies.csv",
    "wincham_leads.csv",
    "wincham_only_companies.csv",
    "hmrc_sic_misclassification.csv",
    "migrated_companies.csv",
    "wincham_sample_data_REDACTED.csv",
]

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
ALT_ROW_FILL  = PatternFill("solid", fgColor="DCE6F1")   # light blue
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT     = Font(name="Calibri", size=10)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN    = Alignment(vertical="center", wrap_text=False)

THIN = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_worksheet(ws, df):
    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    # Data rows
    for row_idx in range(2, len(df) + 2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = BORDER
            if fill:
                cell.fill = fill

    # Auto-fit column widths (capped at 60)
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        try:
            data_max = df.iloc[:, col_idx - 1].fillna("").astype(str).map(len).max() if len(df) > 0 else 0
            if pd.isna(data_max):
                data_max = 0
        except Exception:
            data_max = 0
        max_len = max(len(str(col_name)), int(data_max))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def convert(csv_name):
    csv_path  = os.path.join(BASE_DIR, csv_name)
    xlsx_name = os.path.splitext(csv_name)[0] + ".xlsx"
    xlsx_path = os.path.join(BASE_DIR, xlsx_name)

    print(f"  Converting: {csv_name}")

    try:
        df = pd.read_csv(csv_path, encoding="utf-8-sig", low_memory=False)
    except UnicodeDecodeError:
        df = pd.read_csv(csv_path, encoding="latin-1", low_memory=False)

    row_count = len(df)
    col_count = len(df.columns)

    # Write raw data via pandas first
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")

    # Re-open and apply styling
    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.title = "Data"
    style_worksheet(ws, df)

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Source File"
    ws_sum["B1"] = "Rows"
    ws_sum["C1"] = "Columns"
    ws_sum["D1"] = "Generated"

    for cell in ws_sum[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    from datetime import datetime
    ws_sum["A2"] = csv_name
    ws_sum["B2"] = row_count
    ws_sum["C2"] = col_count
    ws_sum["D2"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    for cell in ws_sum[2]:
        cell.font = DATA_FONT
        cell.border = BORDER

    ws_sum.column_dimensions["A"].width = 45
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 20

    wb.save(xlsx_path)
    print(f"  ✓ Saved:  {xlsx_name}  ({row_count:,} rows, {col_count} columns)")
    return xlsx_name


def main():
    print("\n=== Wincham CSV → Excel Converter ===\n")
    success = []
    errors  = []

    for csv_file in CSV_FILES:
        csv_path = os.path.join(BASE_DIR, csv_file)
        if not os.path.exists(csv_path):
            print(f"  ⚠ SKIPPED (not found): {csv_file}")
            continue
        try:
            name = convert(csv_file)
            success.append(name)
        except Exception as e:
            print(f"  ✗ ERROR on {csv_file}: {e}")
            errors.append(csv_file)

    print(f"\n=== Complete: {len(success)} converted, {len(errors)} errors ===")
    if errors:
        print("Errors:", errors)
    print("\nOutput files:")
    for f in success:
        print(f"  {os.path.join(BASE_DIR, f)}")


if __name__ == "__main__":
    main()
