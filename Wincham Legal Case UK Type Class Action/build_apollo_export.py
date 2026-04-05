"""
Apollo.io Export Builder
Converts the verified victim list into Apollo.io import format.
Apollo matches on: First Name, Last Name, Company Name, LinkedIn URL.
Also adds empty enrichment columns ready to paste Apollo export back in.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, re

BASE = r"C:\DAD\UK_Lanzarote_Repatriation\Wincham Legal Case UK Type Class Action"
MASTER = os.path.join(BASE, "Wincham_Victim_Database_MASTER_v2.xlsx")
APOLLO_CSV  = os.path.join(BASE, "apollo_import_verified_victims.csv")
APOLLO_XLSX = os.path.join(BASE, "Wincham_Victims_Apollo_Enriched.xlsx")

# ── Load verified victims sheet ────────────────────────────────────────────
print("\n=== Building Apollo.io Export ===\n")
df = pd.read_excel(MASTER, sheet_name="✅ VERIFIED VICTIMS", dtype=str)
df = df.fillna("")

# ── Parse "SURNAME, First Names" → first_name / last_name ─────────────────
def parse_name(raw):
    """'SMITH, John David' → ('John', 'Smith')"""
    if not isinstance(raw, str) or not raw.strip():
        return "", ""
    raw = raw.strip()
    if "," in raw:
        parts = raw.split(",", 1)
        last  = parts[0].strip().title()
        first = parts[1].strip().title()
        # Take only the first given name for Apollo matching
        first = first.split()[0] if first else ""
    else:
        words = raw.split()
        last  = words[-1].title() if words else ""
        first = words[0].title()  if len(words) > 1 else ""
    return first, last

df[["First Name", "Last Name"]] = df["Director Name"].apply(
    lambda x: pd.Series(parse_name(x))
)

# ── Parse address into components ──────────────────────────────────────────
def parse_address(addr):
    """Split address string into street, city, postcode."""
    if not isinstance(addr, str) or not addr.strip():
        return "", "", ""
    parts = [p.strip() for p in addr.split(",") if p.strip()]
    postcode = ""
    # UK postcode pattern
    pc_match = re.search(r'\b[A-Z]{1,2}\d{1,2}[A-Z]?\s?\d[A-Z]{2}\b', addr, re.IGNORECASE)
    if pc_match:
        postcode = pc_match.group().upper()
    city     = parts[-2].strip() if len(parts) >= 2 else ""
    street   = ", ".join(parts[:-2]) if len(parts) > 2 else parts[0] if parts else ""
    return street, city, postcode

df[["Street", "City", "Postcode"]] = df["Director Address"].apply(
    lambda x: pd.Series(parse_address(x))
)

# ── Build Apollo import dataframe ─────────────────────────────────────────
apollo_df = pd.DataFrame({
    # Apollo required / high-match fields
    "First Name":           df["First Name"],
    "Last Name":            df["Last Name"],
    "Title":                "",                        # Apollo will enrich
    "Company":              df["Company Name"],
    "Company Number (CH)":  df["Company Number"],
    "Company Status":       df["Status"],

    # Address (helps Apollo geo-match)
    "Street":               df["Street"],
    "City":                 df["City"],
    "State":                "England",
    "Postal Code":          df["Postcode"],
    "Country":              "United Kingdom",

    # Leave blank — Apollo fills these in on enrichment
    "Email":                "",
    "Email Status":         "",
    "Phone":                "",
    "LinkedIn URL":         "",
    "Apollo ID":            "",

    # Context columns for your own reference
    "Wincham Entity":       df["Scheme Address Used"] if "Scheme Address Used" in df.columns else df.get("Address Type", ""),
    "Source File":          df["Source"] if "Source" in df.columns else df.get("Source File", ""),
    "Full Director Name":   df["Director Name"],
    "Full Home Address":    df["Director Address"],
})

# Remove rows with no usable name
apollo_df = apollo_df[
    (apollo_df["First Name"].str.strip() != "") |
    (apollo_df["Last Name"].str.strip()  != "")
].copy()

# ── Save Apollo import CSV ─────────────────────────────────────────────────
apollo_df.to_csv(APOLLO_CSV, index=False, encoding="utf-8-sig")
print(f"  ✓ Apollo import CSV: {APOLLO_CSV}")
print(f"    Rows: {len(apollo_df):,}")

# ── Save enriched XLSX ─────────────────────────────────────────────────────
with pd.ExcelWriter(APOLLO_XLSX, engine="openpyxl") as writer:
    apollo_df.to_excel(writer, sheet_name="Apollo Import", index=False)

    # Instructions sheet
    instructions = pd.DataFrame({
        "Step": [
            "1 — Import CSV into Apollo.io",
            "2 — Run Apollo enrichment",
            "3 — Export from Apollo",
            "4 — Paste back into this file",
            "5 — Validate emails",
            "6 — Remove bounces",
        ],
        "Action": [
            "Go to apollo.io → Lists → Import → Upload 'apollo_import_verified_victims.csv'",
            "In Apollo: select all contacts → Enrich → this fills Email, Phone, LinkedIn",
            "Apollo: Export → CSV (include all fields)",
            "Paste Apollo export columns: Email, Phone, LinkedIn URL, Apollo ID into the 'Apollo Import' sheet",
            "Filter 'Email Status' column — keep only 'Verified' or 'Likely to engage'",
            "Delete rows where email bounced or status is 'Invalid' before licensing to CMC",
        ],
    })
    instructions.to_excel(writer, sheet_name="HOW TO USE APOLLO", index=False)

# Style the workbook
wb = load_workbook(APOLLO_XLSX)

GREEN  = PatternFill("solid", fgColor="1A5C2A")
GOLD   = PatternFill("solid", fgColor="856404")
EMPTY  = PatternFill("solid", fgColor="FFF9C4")   # yellow highlight for blank cols
ALT    = PatternFill("solid", fgColor="E8F5E9")
H_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
D_FONT = Font(name="Calibri", size=10)
E_FONT = Font(name="Calibri", size=10, italic=True, color="888888")
THIN   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Columns that Apollo will fill in (highlight for user attention)
APOLLO_FILL_COLS = {"Email", "Email Status", "Phone", "LinkedIn URL", "Apollo ID"}

ws = wb["Apollo Import"]
for ci, col in enumerate(apollo_df.columns, 1):
    c = ws.cell(row=1, column=ci)
    c.font  = H_FONT
    c.fill  = EMPTY if col in APOLLO_FILL_COLS else GREEN
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER

for ri in range(2, len(apollo_df)+2):
    fill = ALT if ri % 2 == 0 else None
    for ci in range(1, len(apollo_df.columns)+1):
        c = ws.cell(row=ri, column=ci)
        col_name = apollo_df.columns[ci-1]
        c.border = BORDER
        c.alignment = Alignment(vertical="center")
        if col_name in APOLLO_FILL_COLS:
            c.font = E_FONT
            c.fill = EMPTY
        else:
            c.font = D_FONT
            if fill: c.fill = fill

# Column widths
col_widths = {
    "First Name": 15, "Last Name": 20, "Title": 15, "Company": 40,
    "Company Number (CH)": 18, "Company Status": 15, "Street": 35,
    "City": 18, "State": 12, "Postal Code": 12, "Country": 14,
    "Email": 30, "Email Status": 15, "Phone": 18, "LinkedIn URL": 35,
    "Apollo ID": 20, "Wincham Entity": 35, "Source File": 22,
    "Full Director Name": 28, "Full Home Address": 45,
}
for ci, col in enumerate(apollo_df.columns, 1):
    ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 18)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# Style instructions sheet
ws2 = wb["HOW TO USE APOLLO"]
for ci in range(1, 3):
    c = ws2.cell(row=1, column=ci)
    c.font = H_FONT; c.fill = GOLD; c.border = BORDER
    c.alignment = Alignment(horizontal="center")
for ri in range(2, 8):
    for ci in range(1, 3):
        c = ws2.cell(row=ri, column=ci)
        c.font = D_FONT; c.border = BORDER
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 80

wb.save(APOLLO_XLSX)
print(f"  ✓ Apollo enrichment XLSX: {APOLLO_XLSX}")
print(f"\n  ✅ Ready to import into Apollo.io")
print(f"  📋 {len(apollo_df):,} verified victims — clean, operator-free, email columns ready")
