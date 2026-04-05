import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

BASE = r"C:\DAD\UK_Lanzarote_Repatriation\Wincham Legal Case UK Type Class Action"
OUT  = os.path.join(BASE, "Wincham_Victim_Database_MASTER.xlsx")

# ── Wincham entity address signatures ──────────────────────────────────────
WINCHAM_ADDRESSES = {
    "Greenfield Farm Trading Estate, Congleton, CW12 4TR":  "Wincham Accountants (CW12 4TR — Trading Estate)",
    "Greenfield Farm Industrial Estate, Congleton, CW12 4TR": "Wincham Accountants (CW12 4TR — Industrial Estate)",
    "Greenfield Farm Industrial Estate, Cheshire, CW12 4TR": "Wincham Accountants (CW12 4TR — Industrial Estate)",
    "Greenfields Farm Trading Estate, Congleton, CW12 4TR":  "Wincham Accountants (CW12 4TR — Trading Estate)",
    "Greenfield Industrial Estate, Congleton, CW12 4TR":     "Wincham Accountants (CW12 4TR — Industrial Estate)",
    "Greenfield Farm Trading, Congleton, CW12 4TR":          "Wincham Accountants (CW12 4TR — Trading Estate)",
    "Back Lane, Congleton, CW12 4TR":                        "Wincham Accountants (CW12 4TR)",
    "Back Lane Sunlit House, Congleton, CW12 4TR":           "Wincham Accountants (CW12 4TR)",
    "The Courtyard, Congleton, CW12 4TR":                    "Wincham Accountants (CW12 4TR — The Courtyard)",
    "Wincham House, Congleton, CW12 4TR":                    "Wincham House (Early Wincham Entity)",
    "22 Oak Drive, Scholar Green, ST7 3LY":                  "Malcolm Roach Personal Address",
    "2 Woodberry Grove, Finchley, N12 0DR":                  "Michael Duke / Marion Black (Finchley)",
    "Floor, Manchester, M25 9JY":                            "Yomtov Jacobs (Manchester)",
    "47 Bury New Road, Manchester, M25 9JY":                 "Yomtov Jacobs (Manchester)",
    "Bury New Road, Manchester, M25 9JY":                    "Yomtov Jacobs (Manchester)",
    "Leicester Road, Manchester, M7 4AS":                    "Yomtov Jacobs (Manchester)",
    "Prestwich, Manchester, M25 9JY":                        "Osker Heiman (Manchester)",
}

# Adrem CW12 4AA addresses
ADREM_ADDRESSES = {
    "1-2 Albert Chambers, Canal Street, Congleton, CW12 4AA": "Adrem Accounting Ltd (CW12 4AA)",
    "Canal Street, Congleton, CW12 4AA":                       "Adrem Accounting Ltd (CW12 4AA)",
}

ALL_SCHEME_SIGS = list(WINCHAM_ADDRESSES.keys()) + list(ADREM_ADDRESSES.keys())

def is_scheme_address(addr):
    if not isinstance(addr, str):
        return False
    addr_lower = addr.lower()
    for sig in ALL_SCHEME_SIGS:
        if sig.lower() in addr_lower:
            return True
    return False

def entity_label(addr):
    if not isinstance(addr, str):
        return "Unknown"
    addr_lower = addr.lower()
    for sig, label in ADREM_ADDRESSES.items():
        if sig.lower() in addr_lower:
            return label
    for sig, label in WINCHAM_ADDRESSES.items():
        if sig.lower() in addr_lower:
            return label
    return "Other / Unknown"

# ── Load and prep the two main datasets ───────────────────────────────────
def load_leads(fname, source_label):
    path = os.path.join(BASE, fname)
    df = pd.read_csv(path, encoding="utf-8-sig", low_memory=False)
    # Standardise column names
    df.columns = df.columns.str.strip()
    # Both files have: Company Name, Company Number, Status, Date of Creation, Director Name, Director Address (LEAD)
    df = df.rename(columns={
        "Director Address (LEAD)": "Director Address",
        "Date of Creation": "Date Incorporated",
    })
    df["Source File"] = source_label
    return df

wincham_df = load_leads("wincham_leads.csv",       "wincham_leads.csv")
adrem_df   = load_leads("adrem_leads.csv",         "adrem_leads.csv")
adrem_cw   = pd.read_csv(os.path.join(BASE, "adrem_cw12_4aa_leads.csv"), encoding="utf-8-sig", low_memory=False)
adrem_cw.columns = adrem_cw.columns.str.strip()
adrem_cw = adrem_cw.rename(columns={
    "Director Correspondence Address (LEAD)": "Director Address",
    "Date Incorporated": "Date Incorporated",
})
adrem_cw["Source File"] = "adrem_cw12_4aa_leads.csv"

# Pool all data together
all_df = pd.concat([wincham_df, adrem_df, adrem_cw], ignore_index=True)

# Ensure essential columns exist
for col in ["Company Name","Company Number","Status","Date Incorporated","Director Name","Director Address","Source File"]:
    if col not in all_df.columns:
        all_df[col] = ""

# ── Classify each row ─────────────────────────────────────────────────────
all_df["Scheme Address Used"]   = all_df["Director Address"].apply(entity_label)
all_df["Is Scheme Address"]     = all_df["Director Address"].apply(is_scheme_address)

# Real victims = rows where director address is NOT a known scheme address
victims_df = all_df[~all_df["Is Scheme Address"]].copy()

# Operator rows (Roach, Duke, Jacobs etc.) = scheme address rows
operators_df = all_df[all_df["Is Scheme Address"]].copy()

# ── Build output columns for victim sheet ─────────────────────────────────
VICTIM_COLS = [
    "Director Name",
    "Director Address",       # real home address
    "Company Name",
    "Company Number",
    "Status",
    "Date Incorporated",
    "Scheme Address Used",    # which Wincham entity address was on the company
    "Source File",
]

# Entity groupings for separate sheets
entity_groups = {
    "Wincham CW12 4TR (All)":    all_df[all_df["Scheme Address Used"].str.contains("Wincham Accountants", na=False)],
    "Wincham House":             all_df[all_df["Scheme Address Used"].str.contains("Wincham House", na=False)],
    "Adrem CW12 4AA":            all_df[all_df["Scheme Address Used"].str.contains("Adrem", na=False)],
    "Roach Personal (ST7)":      all_df[all_df["Scheme Address Used"].str.contains("Malcolm Roach", na=False)],
    "Duke-Black Finchley (N12)": all_df[all_df["Scheme Address Used"].str.contains("Duke", na=False)],
    "Jacobs Manchester (M25)":   all_df[all_df["Scheme Address Used"].str.contains("Jacobs", na=False)],
}

# ── Styling helpers ────────────────────────────────────────────────────────
NAVY   = PatternFill("solid", fgColor="1F3864")
RED    = PatternFill("solid", fgColor="C00000")
GOLD   = PatternFill("solid", fgColor="C9A800")
TEAL   = PatternFill("solid", fgColor="1E6B7A")
LAVEND = PatternFill("solid", fgColor="4B3A8C")
DARK   = PatternFill("solid", fgColor="2D2D2D")
ALT    = PatternFill("solid", fgColor="EBF3FB")
ALT2   = PatternFill("solid", fgColor="FFF2CC")

H_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
D_FONT = Font(name="Calibri", size=10)
THIN   = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILLS = [NAVY, RED, GOLD, TEAL, LAVEND, DARK]

def style_sheet(ws, df, header_fill=None):
    hf = header_fill or NAVY
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci)
        c.value = col
        c.font  = H_FONT
        c.fill  = hf
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    alt = ALT
    for ri in range(2, len(df)+2):
        fill = alt if ri % 2 == 0 else None
        for ci in range(1, len(df.columns)+1):
            c = ws.cell(row=ri, column=ci)
            c.font   = D_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical="center")
            if fill:
                c.fill = fill

    for ci, col in enumerate(df.columns, 1):
        try:
            data_max = df.iloc[:, ci-1].fillna("").astype(str).map(len).max() if len(df) > 0 else 0
            if pd.isna(data_max): data_max = 0
        except:
            data_max = 0
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(str(col)), int(data_max)) + 4, 70)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ── Write workbook ─────────────────────────────────────────────────────────
print(f"\n=== Building Wincham Victim Database Master Workbook ===\n")

# First write with pandas
sheets_data = {
    "ALL VICTIMS (Home Addresses)": victims_df[VICTIM_COLS].drop_duplicates().sort_values("Director Name"),
    "ALL OPERATORS": operators_df[VICTIM_COLS].drop_duplicates().sort_values("Scheme Address Used"),
}
for name, grp in entity_groups.items():
    clean = grp[[c for c in VICTIM_COLS if c in grp.columns]].drop_duplicates()
    sheets_data[name[:31]] = clean  # Excel sheet name max 31 chars

with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    for sheet_name, data in sheets_data.items():
        data.to_excel(writer, sheet_name=sheet_name, index=False)

# Re-open and style
wb = load_workbook(OUT)
fills_cycle = [NAVY, GOLD, RED, TEAL, LAVEND, DARK, NAVY, GOLD, RED]

for i, sheet_name in enumerate(sheets_data.keys()):
    ws = wb[sheet_name]
    style_sheet(ws, sheets_data[sheet_name], header_fill=fills_cycle[i % len(fills_cycle)])
    print(f"  ✓ Sheet: '{sheet_name}'  ({len(sheets_data[sheet_name]):,} rows)")

# Summary sheet
ws_sum = wb.create_sheet("SUMMARY")
ws_sum["A1"] = "Sheet Name"
ws_sum["B1"] = "Row Count"
ws_sum["C1"] = "Description"
ws_sum["D1"] = "Generated"

for c in ws_sum[1]:
    c.font = H_FONT
    c.fill = NAVY
    c.border = BORDER
    c.alignment = Alignment(horizontal="center")

descs = {
    "ALL VICTIMS (Home Addresses)": "Directors whose address is NOT a Wincham/Adrem scheme address — potential claimants",
    "ALL OPERATORS":                "Directors using a scheme address — Wincham staff/nominees",
    "Wincham CW12 4TR (All)":      "All companies registered through Wincham CW12 4TR addresses",
    "Wincham House":                "Early-era Wincham House entity companies",
    "Adrem CW12 4AA":               "Companies serviced through Adrem Accounting LLC (CW12 4AA)",
    "Roach Personal (ST7)":         "Malcolm Roach's personal address used as director address",
    "Duke-Black Finchley (N12)":    "Michael Duke / Marion Black (Finchley N12) – nominee directors",
    "Jacobs Manchester (M25)":      "Yomtov Jacobs (Manchester M25) – nominee director",
}

for ri, (sheet_name, data) in enumerate(sheets_data.items(), 2):
    ws_sum.cell(row=ri, column=1).value = sheet_name
    ws_sum.cell(row=ri, column=2).value = len(data)
    ws_sum.cell(row=ri, column=3).value = descs.get(sheet_name, "")
    ws_sum.cell(row=ri, column=4).value = datetime.now().strftime("%Y-%m-%d %H:%M")
    for ci in range(1, 5):
        c = ws_sum.cell(row=ri, column=ci)
        c.font = D_FONT
        c.border = BORDER
        if ri % 2 == 0:
            c.fill = ALT

ws_sum.column_dimensions["A"].width = 35
ws_sum.column_dimensions["B"].width = 12
ws_sum.column_dimensions["C"].width = 70
ws_sum.column_dimensions["D"].width = 20
ws_sum.freeze_panes = "A2"

# Move SUMMARY to first position
wb.move_sheet("SUMMARY", offset=-len(wb.sheetnames)+1)

wb.save(OUT)
print(f"\n✓ Master workbook saved: {OUT}")
print(f"  Total sheets: {len(wb.sheetnames)}")
