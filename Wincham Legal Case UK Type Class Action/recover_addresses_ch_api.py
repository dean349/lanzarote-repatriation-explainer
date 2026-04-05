"""
Companies House API — Real Address Recovery
For victims whose address was replaced by a Wincham scheme address,
look up their ACTUAL correspondence address via the CH Officers API.
Rate-limited to stay within CH free-tier limits (600 req/5 min).
"""

import pandas as pd
import requests, time, os, json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE    = r"C:\DAD\UK_Lanzarote_Repatriation\Wincham Legal Case UK Type Class Action"
MASTER  = os.path.join(BASE, "Wincham_Victim_Database_MASTER_v2.xlsx")
OUT_CSV = os.path.join(BASE, "scheme_addr_recovered.csv")
OUT_XLS = os.path.join(BASE, "scheme_addr_recovered.xlsx")
LOG_F   = os.path.join(BASE, "ch_lookup_log.json")

# ── CH API config ─────────────────────────────────────────────────────────
# Your existing Companies House API key
API_KEY  = "4416608c-08ec-449e-8c5a-ddc66b5bb6b3"
BASE_URL = "https://api.company-information.service.gov.uk"
RATE_LIMIT_PAUSE = 0.55   # ~108 req/min — well inside 600/5min limit

SCHEME_KEYWORDS = [
    "CW12 4TR", "CW12 4AA", "Wincham House",
    "Oak Drive, Scholar Green", "Woodberry Grove", "M25 9JY", "M7 4AS"
]

def is_scheme(addr):
    if not isinstance(addr, str): return True
    return any(kw.lower() in addr.lower() for kw in SCHEME_KEYWORDS)

def get_officers(company_number):
    """Fetch all officers for a company from CH API."""
    cn = str(company_number).strip().zfill(8)
    url = f"{BASE_URL}/company/{cn}/officers"
    try:
        r = requests.get(url, auth=(API_KEY, ""), timeout=10)
        if r.status_code == 200:
            return r.json().get("items", [])
        elif r.status_code == 429:
            print("    ⚠ Rate limited — waiting 10s...")
            time.sleep(10)
            r2 = requests.get(url, auth=(API_KEY, ""), timeout=10)
            if r2.status_code == 200:
                return r2.json().get("items", [])
    except Exception as e:
        print(f"    Error: {e}")
    return []

def extract_address(officer_item):
    """Pull the best available address from a CH officer dict."""
    # Try correspondence address first, then usual residential address
    for addr_key in ["correspondence_address", "address"]:
        addr = officer_item.get(addr_key, {})
        if not addr:
            continue
        parts = []
        for field in ["premises", "address_line_1", "address_line_2",
                      "locality", "region", "postal_code", "country"]:
            val = addr.get(field, "").strip()
            if val:
                parts.append(val)
        if parts:
            return ", ".join(parts)
    return ""

# ── Load scheme address entries ───────────────────────────────────────────
print("\n=== Companies House Address Recovery ===\n")
df = pd.read_excel(MASTER, sheet_name="📋 Scheme Address Entries", dtype=str)
df = df.fillna("")

# Keep unique company + director combinations
df["_key"] = df["Company Number"].str.strip() + "|" + df["Director Name"].str.strip()
df = df.drop_duplicates(subset="_key")
print(f"  Unique company-director pairs to look up: {len(df):,}")

# Load existing cache if any
cache = {}
if os.path.exists(LOG_F):
    with open(LOG_F) as f:
        cache = json.load(f)
    print(f"  Cache loaded: {len(cache):,} companies already fetched")

results = []
companies_fetched = set()

for idx, row in df.iterrows():
    company_num    = str(row.get("Company Number", "")).strip()
    director_name  = str(row.get("Director Name", "")).strip()
    current_addr   = str(row.get("Director Address", "")).strip()

    if not company_num or company_num == "nan":
        continue

    # Use cache if available
    if company_num not in cache:
        if company_num not in companies_fetched:
            officers = get_officers(company_num)
            cache[company_num] = officers
            companies_fetched.add(company_num)
            time.sleep(RATE_LIMIT_PAUSE)
        else:
            officers = cache.get(company_num, [])
    else:
        officers = cache[company_num]

    # Match officer by name (flexible)
    recovered_addr = ""
    matched_officer = None
    dir_surname = director_name.split(",")[0].strip().upper() if "," in director_name else director_name.upper()

    for officer in officers:
        officer_name = officer.get("name", "").upper()
        if dir_surname and dir_surname in officer_name:
            addr = extract_address(officer)
            if addr and not is_scheme(addr):
                recovered_addr = addr
                matched_officer = officer.get("name", "")
                break

    results.append({
        "Director Name":         director_name,
        "Original Scheme Addr":  current_addr,
        "Recovered Home Address": recovered_addr,
        "CH Name Match":         matched_officer or "",
        "Address Recovered":     "✅ Yes" if recovered_addr else "❌ No",
        "Company Name":          row.get("Company Name", ""),
        "Company Number":        company_num,
        "Status":                row.get("Status", ""),
        "Date Incorporated":     row.get("Date Incorporated", ""),
        "Source":                row.get("Source", row.get("Source File", "")),
    })

    total_done = idx + 1
    if total_done % 50 == 0:
        # Save cache checkpoint
        with open(LOG_F, "w") as f:
            json.dump(cache, f)
        recovered_so_far = sum(1 for r in results if r["Address Recovered"] == "✅ Yes")
        print(f"  Progress: {total_done}/{len(df)}  |  Recovered: {recovered_so_far}")

# Final cache save
with open(LOG_F, "w") as f:
    json.dump(cache, f)

# ── Build output ──────────────────────────────────────────────────────────
out_df = pd.DataFrame(results)
recovered = out_df[out_df["Address Recovered"] == "✅ Yes"]
not_recovered = out_df[out_df["Address Recovered"] == "❌ No"]

out_df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

print(f"\n  ✅ Total processed : {len(out_df):,}")
print(f"  ✅ Addresses recovered : {len(recovered):,}")
print(f"  ❌ Could not recover  : {len(not_recovered):,}")

# Save to XLSX
with pd.ExcelWriter(OUT_XLS, engine="openpyxl") as writer:
    recovered.to_excel(writer,     sheet_name="✅ Recovered Addresses", index=False)
    not_recovered.to_excel(writer, sheet_name="❌ Not Recovered",        index=False)
    out_df.to_excel(writer,        sheet_name="All Results",             index=False)

print(f"\n  Saved: {OUT_XLS}")
print(f"\n  Next step: merge 'recovered.csv' into master victim list")
print(f"  Run: python merge_recovered_into_master.py")
